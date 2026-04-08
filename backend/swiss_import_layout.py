"""
Plantillas e importación layout Excel/CSV para catálogo maestro (productos) y métodos de pago.
"""

from __future__ import annotations

import csv
import io
import re
import uuid
from typing import Any, BinaryIO, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from sqlalchemy.orm import Session

from . import models


def _norm_suc_key(n: str) -> str:
    x = n.upper().replace("_", " ")
    x = re.sub(r"\s+", " ", x).strip()
    return x


def build_product_template_xlsx(sucursales: List[models.Sucursal]) -> bytes:
    """Fila 1: MASTER + nombre sucursal en B,D,F,... Fila 2: DESCRIPCION_DASHBOARD + pares CODIGO/DESCRIPCION."""
    wb = Workbook()
    ws = wb.active
    ws.title = "catalogo"
    ws["A1"] = "MASTER"
    col = 2
    for s in sucursales:
        ws.cell(row=1, column=col, value=s.nombre)
        ws.cell(row=1, column=col + 1, value="")
        col += 2
    ws["A2"] = "DESCRIPCION_DASHBOARD"
    col = 2
    for _ in sucursales:
        ws.cell(row=2, column=col, value="CODIGO")
        ws.cell(row=2, column=col + 1, value="DESCRIPCION")
        col += 2
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_metodos_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "metodos"
    ws["A1"] = "NOMBRE_CANONICO"
    ws["B1"] = "ALIAS_POS"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sucursal_lookup_maps(db: Session) -> tuple[dict[str, str], dict[str, str]]:
    rows = db.query(models.Sucursal).all()
    exact: dict[str, str] = {}
    fuzzy: dict[str, str] = {}
    for s in rows:
        n = (s.nombre or "").strip()
        if not n:
            continue
        exact[n] = s.id
        key = _norm_suc_key(n)
        fuzzy.setdefault(key, s.id)
    return exact, fuzzy


def _resolve_sucursal_id(header: str, exact: dict[str, str], fuzzy: dict[str, str]) -> Optional[str]:
    h = (header or "").strip()
    if not h:
        return None
    if h in exact:
        return exact[h]
    return fuzzy.get(_norm_suc_key(h))


def parse_product_layout_from_xlsx(
    data: BinaryIO,
    *,
    db: Session,
    catalogo_id: str,
    allowed_sucursal_ids: set[str],
) -> dict[str, Any]:
    wb = load_workbook(data, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    row0 = next(rows_iter, None)
    row1 = next(rows_iter, None)
    if not row0 or not row1:
        wb.close()
        return {"ok": False, "error": "Archivo vacío o sin encabezados", "rules_applied": 0, "errors": []}

    r0 = list(row0)
    r1 = list(row1)

    if not r0 or len(r0) < 3:
        wb.close()
        return {"ok": False, "error": "Fila 1 inválida (se espera MASTER y bloques por sucursal)", "rules_applied": 0, "errors": []}

    exact, fuzzy = _sucursal_lookup_maps(db)
    blocks: List[Tuple[int, str]] = []
    col = 1
    max_col = max(len(r0), len(r1))
    while col < max_col:
        h = r0[col] if col < len(r0) else None
        name = str(h).strip() if h is not None and str(h).strip() else ""
        cod_lbl = str(r1[col]).strip().upper() if col < len(r1) and r1[col] is not None else ""
        desc_lbl = str(r1[col + 1]).strip().upper() if col + 1 < len(r1) and r1[col + 1] is not None else ""
        if not name:
            col += 2
            continue
        sid = _resolve_sucursal_id(name, exact, fuzzy)
        if not sid:
            wb.close()
            return {
                "ok": False,
                "error": f"Sucursal no reconocida en columna {col + 1}: {name!r}",
                "rules_applied": 0,
                "errors": [],
            }
        if sid not in allowed_sucursal_ids:
            wb.close()
            return {
                "ok": False,
                "error": f"La sucursal {name!r} no está vinculada a este catálogo",
                "rules_applied": 0,
                "errors": [],
            }
        if cod_lbl != "CODIGO" or desc_lbl != "DESCRIPCION":
            wb.close()
            return {
                "ok": False,
                "error": f"Fila 2: se esperaba CODIGO/DESCRIPCION bajo {name!r} (cols {col + 1}-{col + 2})",
                "rules_applied": 0,
                "errors": [],
            }
        blocks.append((col, sid))
        col += 2

    if not blocks:
        wb.close()
        return {"ok": False, "error": "No se encontraron bloques de sucursal válidos", "rules_applied": 0, "errors": []}

    errors: List[str] = []
    rules: List[Tuple[str, str]] = []
    row_num = 3
    for data_row in rows_iter:
        r = list(data_row)
        if not r or all(v is None or str(v).strip() == "" for v in r):
            row_num += 1
            continue
        maestro = str(r[0]).strip() if len(r) > 0 and r[0] is not None else ""
        if not maestro:
            errors.append(f"Fila {row_num}: columna A (maestro) vacía, se omite")
            row_num += 1
            continue
        for start_col, _sid in blocks:
            cod = str(r[start_col]).strip() if start_col < len(r) and r[start_col] is not None else ""
            desc = str(r[start_col + 1]).strip() if start_col + 1 < len(r) and r[start_col + 1] is not None else ""
            if not cod and not desc:
                continue
            if not desc:
                errors.append(f"Fila {row_num}: falta DESCRIPCION para sucursal en cols {start_col + 1}-{start_col + 2}")
                continue
            alias_local = desc.upper().strip()
            if not alias_local:
                continue
            rules.append((maestro, alias_local))
        row_num += 1

    wb.close()

    applied = 0
    for maestro, alias_local in rules:
        db.query(models.CatalogoMaestroProducto).filter(
            models.CatalogoMaestroProducto.catalogo_id == catalogo_id,
            models.CatalogoMaestroProducto.alias_local == alias_local,
        ).delete(synchronize_session=False)
        db.add(
            models.CatalogoMaestroProducto(
                catalogo_id=catalogo_id,
                nombre_maestro=maestro,
                alias_local=alias_local,
            )
        )
        applied += 1
    db.commit()

    return {"ok": True, "error": None, "rules_applied": applied, "errors": errors[:200]}


def parse_product_layout_from_csv(
    raw: bytes,
    *,
    db: Session,
    catalogo_id: str,
    allowed_sucursal_ids: set[str],
) -> dict[str, Any]:
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if len(rows) < 3:
        return {"ok": False, "error": "CSV: mínimo 3 filas (encabezados + datos)", "rules_applied": 0, "errors": []}

    def pad(row: List[str], n: int) -> List[str]:
        return list(row) + [""] * max(0, n - len(row))

    r0 = pad(rows[0], 256)
    r1 = pad(rows[1], 256)
    exact, fuzzy = _sucursal_lookup_maps(db)
    blocks: List[Tuple[int, str]] = []
    col = 1
    max_col = max(len(r0), len(r1))
    while col < max_col:
        name = r0[col].strip() if col < len(r0) else ""
        cod_lbl = r1[col].strip().upper() if col < len(r1) else ""
        desc_lbl = r1[col + 1].strip().upper() if col + 1 < len(r1) else ""
        if not name:
            col += 2
            continue
        sid = _resolve_sucursal_id(name, exact, fuzzy)
        if not sid:
            return {"ok": False, "error": f"Sucursal no reconocida: {name!r}", "rules_applied": 0, "errors": []}
        if sid not in allowed_sucursal_ids:
            return {"ok": False, "error": f"Sucursal {name!r} no vinculada al catálogo", "rules_applied": 0, "errors": []}
        if cod_lbl != "CODIGO" or desc_lbl != "DESCRIPCION":
            return {"ok": False, "error": f"Fila 2 CSV inválida bajo {name!r}", "rules_applied": 0, "errors": []}
        blocks.append((col, sid))
        col += 2

    if not blocks:
        return {"ok": False, "error": "Sin bloques de sucursal", "rules_applied": 0, "errors": []}

    errors: List[str] = []
    rules: List[Tuple[str, str]] = []
    row_num = 3
    for data_row in rows[2:]:
        r = pad(data_row, max_col + 2)
        if all(not x.strip() for x in r):
            row_num += 1
            continue
        maestro = r[0].strip()
        if not maestro:
            errors.append(f"Fila {row_num}: maestro vacío")
            row_num += 1
            continue
        for start_col, _sid in blocks:
            cod = r[start_col].strip() if start_col < len(r) else ""
            desc = r[start_col + 1].strip() if start_col + 1 < len(r) else ""
            if not cod and not desc:
                continue
            if not desc:
                errors.append(f"Fila {row_num}: falta descripción")
                continue
            alias_local = desc.upper().strip()
            if alias_local:
                rules.append((maestro, alias_local))
        row_num += 1

    applied = 0
    for maestro, alias_local in rules:
        db.query(models.CatalogoMaestroProducto).filter(
            models.CatalogoMaestroProducto.catalogo_id == catalogo_id,
            models.CatalogoMaestroProducto.alias_local == alias_local,
        ).delete(synchronize_session=False)
        db.add(
            models.CatalogoMaestroProducto(
                catalogo_id=catalogo_id,
                nombre_maestro=maestro,
                alias_local=alias_local,
            )
        )
        applied += 1
    db.commit()

    return {"ok": True, "error": None, "rules_applied": applied, "errors": errors[:200]}


def parse_metodos_from_xlsx(
    data: BinaryIO,
    *,
    db: Session,
    sucursal_id: str,
    norm_fn,
) -> dict[str, Any]:
    suc = db.query(models.Sucursal).filter(models.Sucursal.id == sucursal_id).first()
    if not suc:
        return {"ok": False, "error": "Sucursal no encontrada", "rules_applied": 0, "errors": []}

    wb = load_workbook(data, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(min_row=2, values_only=True)
    errors: List[str] = []
    applied = 0
    row_num = 2
    for row in rows_iter:
        row_num += 1
        if not row or (row[0] is None and (len(row) < 2 or row[1] is None)):
            continue
        canon = str(row[0]).strip() if row[0] is not None else ""
        alias = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if not canon or not alias:
            if canon or alias:
                errors.append(f"Fila {row_num}: canon y alias requeridos")
            continue
        alias_norm = norm_fn(alias)
        if not alias_norm:
            errors.append(f"Fila {row_num}: alias inválido")
            continue
        existing = (
            db.query(models.MetodoPagoAlias)
            .filter(
                models.MetodoPagoAlias.sucursal_id == sucursal_id,
                models.MetodoPagoAlias.alias_norm == alias_norm,
            )
            .first()
        )
        if existing:
            existing.alias = alias
            existing.nombre_canonico = canon
        else:
            db.add(
                models.MetodoPagoAlias(
                    id=str(uuid.uuid4()),
                    sucursal_id=sucursal_id,
                    alias=alias,
                    alias_norm=alias_norm,
                    nombre_canonico=canon,
                )
            )
        applied += 1
    wb.close()
    db.commit()
    return {"ok": True, "error": None, "rules_applied": applied, "errors": errors[:200]}


def parse_metodos_from_csv(
    raw: bytes,
    *,
    db: Session,
    sucursal_id: str,
    norm_fn,
) -> dict[str, Any]:
    suc = db.query(models.Sucursal).filter(models.Sucursal.id == sucursal_id).first()
    if not suc:
        return {"ok": False, "error": "Sucursal no encontrada", "rules_applied": 0, "errors": []}

    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    next(reader, None)
    errors: List[str] = []
    applied = 0
    row_num = 1
    for row in reader:
        row_num += 1
        if not row or len(row) < 2:
            continue
        canon = row[0].strip()
        alias = row[1].strip()
        if not canon or not alias:
            continue
        alias_norm = norm_fn(alias)
        if not alias_norm:
            errors.append(f"Fila {row_num}: alias inválido")
            continue
        existing = (
            db.query(models.MetodoPagoAlias)
            .filter(
                models.MetodoPagoAlias.sucursal_id == sucursal_id,
                models.MetodoPagoAlias.alias_norm == alias_norm,
            )
            .first()
        )
        if existing:
            existing.alias = alias
            existing.nombre_canonico = canon
        else:
            db.add(
                models.MetodoPagoAlias(
                    id=str(uuid.uuid4()),
                    sucursal_id=sucursal_id,
                    alias=alias,
                    alias_norm=alias_norm,
                    nombre_canonico=canon,
                )
            )
        applied += 1
    db.commit()
    return {"ok": True, "error": None, "rules_applied": applied, "errors": errors[:200]}
